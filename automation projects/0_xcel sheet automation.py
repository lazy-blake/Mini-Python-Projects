import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

wb = xl.load_workbook(
    "C:/Users/akash/OneDrive/Documents/Python/Projects/automation projects/transaction.xlsx"
)
sheet = wb["Sheet1"]
# cell = sheet["b1"]
# cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    price = sheet.cell(row, 3)
    new_price = price.value * 0.9
    sheet.cell(row, 4).value = new_price

# to add a chart
value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(value)
chart.title = "New Updated Price chart"

# to add colors
series = chart.series[0]
fill = PatternFillProperties(prst="pct5")
fill.foreground = ColorChoice(prstClr="red")
fill.background = ColorChoice(prstClr="blue")
series.graphicalProperties.pattFill = fill


sheet.add_chart(chart, "b8")

wb.save(
    "C:/Users/akash/OneDrive/Documents/Python/Projects/automation projects/transaction2.xlsx"
)
